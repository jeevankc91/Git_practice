import os
import openpyxl

os.chdir('F:\git_practice_dev1\Git_practice')
wb = openpyxl.load_workbook('0WDAGAPP_F130D11_PIT_REPORT.xlsx')

Sheet_name = wb.get_sheet_names()
number_of_sheets = len(Sheet_name)
my_sheet = 0
for i in range(number_of_sheets):
    if (Sheet_name[i] == 'UNUSED_VARIABLES'):
        my_sheet = i
sheet = wb.get_sheet_by_name(Sheet_name[my_sheet])
row_count = sheet.max_row
column_count = sheet.max_column
print row_count,column_count
print sheet.cell(row=1,column=1).value
